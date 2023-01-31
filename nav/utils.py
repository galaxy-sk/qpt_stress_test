import os
import pyodbc
import datetime as dt
import numpy as np
import pandas as pd
#import xlsxwriter
from pandas.api.types import CategoricalDtype
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
pd.options.display.max_rows = None
pd.options.display.float_format = '{:,.2f}'.format
#from sendgrid import SendGridAPIClient
#from sendgrid.helpers.mail import Mail

def _get_data(qry):
    conn = pyodbc.connect("DRIVER={SQL Server};SERVER=10.17.181.234;PORT=1433;UID=ro;PWD=ro;DATABASE=Trading", autocommit=False)
    results = pd.read_sql_query(qry, conn).fillna(0)
    conn.close()
    return results

def get_marks(daytime):
    df = _pull_sql_marks(daytime)
    df2 = _pull_sql_marks_2(daytime)
    df = _format_marks(df, df2, daytime)
    return df

def _pull_sql_marks(daytime):
    qry = """select format(timestamp,'yyyyMMdd') as DATE, symbol as Currency, _close as Mark, 'coinmarketcap' as Source
    from RawData.cnmk_1_m.ohlcv_historical
    where format(timestamp,'yyyyMMdd') = '{int_day}' and
    name not in ('CyberMiles','Genesis Mana','NFT','NEFTiPEDiA',
    'UNICORN Token','DEONEX COIN','Don-key','LOLTOKEN','SoMee.Social [OLD]')"""
    qry = qry.format(qry, int_day=daytime.strftime('%Y%m%d'))
    results = _get_data(qry)
    return results

# included as cmc table doesn't have GBP and EUR
def _pull_sql_marks_2(daytime):
    qry = "select DATE, Currency, Mark, Source from trading.pnl.marks where DATE = {int_day}"
    qry = qry.format(qry, int_day=daytime.strftime('%Y%m%d'))               
    results = _get_data(qry)
    return results

def _format_marks(df, df2, daytime):
    df = df.append(df2, ignore_index=True)
    df.Currency = df.Currency.apply(str.upper)
    df.drop(df[df.Currency.duplicated(keep="first")].index, axis=0, inplace=True)
    df.set_index('Currency', inplace=True)
    # adding coins which are having a different name but essentially same as an existing coin
    df.loc['AUSDC'] = df.loc['USDC'].values
    df.loc['AWETH'] = df.loc['WETH'].values
    df.loc['FRXETH'] = df.loc['ETH'].values
    return df

def get_assets(daytime, exchanges, marks):
    df = _pull_sql_assets(daytime, exchanges)
    df = _format_assets(daytime, df, marks)
    return df

def _pull_sql_assets(daytime, exchanges):
    qry = """select a.Account, a.Balance, a.BalanceType, UPPER(a.Currency) as Currency, 
    b.TableName as Source, FORMAT(a.Date_UTC,'yyyyMMdd') as Timestamp, a.AsOf_UTC as Timestamp_Native
    from Operations.balances.EndOfDay_00UTC A LEFT JOIN Operations.balances.sources B on a.Account = b.Account
    where balance != 0 and a.Date_UTC = '{trade_date}' and Currency not like '%SWAP%' and a.Account in {exchanges} and Seconds_from_00UTC <= 79259
    order by Account""".format(trade_date = daytime, exchanges = tuple(exchanges))
    df = _get_data(qry)
    return df

def _format_assets(daytime, df, marks):
#     df.drop(df.loc[(df.Account.isin(['OKEX-2-W1', 'OKEX-2-U1', 'OKEX-2-S1', 'OKEX-2-S2', 'OKEX-2-S3', 'HUBI-1-M-P', 'FUBI-M','FUB2-M','HUBI-3-S3-F'])) 
#                    & (df.BalanceType!='Balance')].index, axis=0, inplace=True) #comment out to include unrealized
    df.drop(df.loc[(df.Account.isin(['HUBI-M','HUB2-M'])) & (df.BalanceType.isin(['Margin Loan','Unrealized']))].index, axis=0, inplace=True)
    df.Account.replace(['OKEX-2-W1', 'OKEX-2-U1', 'FBLK-AUDIT-BINE', 'FBLK-AUDIT-HUBI', 'FBLK-AUDIT-OKEX', 'FBLK-Default', 'FBLK-GOTC-GACM', 'FBLK-LEND-BTGO', 'FBLK-LEND-DRAW', 
                        'FBLK-LEND-OXTF', 'FBLK-LEND-XRPF', 'FBLK-LMAC-M', 'FBLK-Network Deposits', 'FBLK-PITX-E', 'FBLK-LEND-CELS', 'FBLK-LEND-GADI', 'FBLK-LEND-NICO', 
                        'FBLK-LEND-GENX','FBLK-BINE-MX-S1','FBLK-DEFI-AAVE','FBLK-GOTC-BIGO','FBLK-DYDX-1-M-P','FBLK-WOOX-1-M-E','FBLK-LEND-GACM'], 
                       ['OKEX-2-M-W', 'OKEX-2-M', 'FBLK', 'FBLK', 'FBLK', 'FBLK', 'FBLK', 'FBLK', 'FBLK', 'FBLK', 'FBLK', 'FBLK', 'FBLK', 'FBLK','FBLK', 'FBLK', 'FBLK', 'FBLK', 'FBLK', 'FBLK', 
                        'FBLK', 'FBLK', 'FBLK', 'FBLK'], inplace=True)

    df.drop(df.loc[(df.Account.isin(['DEFI-STRAT-8'])) & (df.BalanceType.isin(['wallet','borrow_token']))].index, axis=0, inplace=True)
    df.drop(df.loc[(df.Account.isin(['WOOX-1-M-E'])) & (df.Balance<0) & (~df.BalanceType.isin(['Unrealized']))].index, axis=0, inplace=True)
    df.drop(df.loc[(df.Account.isin(['FTXE-1-M-E'])) & (df.Balance<0) & (~df.BalanceType.isin(['Unrealized']))].index, axis=0, inplace=True)
    df.drop(df.loc[(df.Account.isin(['OKEX-2-M'])) & (df.Balance<0) & (~df.BalanceType.isin(['Unrealized']))].index, axis=0, inplace=True)
    df.drop(df.loc[(df.Account.isin(['OKEX-2-S3'])) & (df.Balance<0) & (~df.BalanceType.isin(['Unrealized']))].index, axis=0, inplace=True)
    df.drop(df.loc[(df.Account.isin(['BINE-2-S1-M','BINE-2-S2-M'])) & (df.Balance<0) & (~df.BalanceType.isin(['Unrealized']))].index, axis=0, inplace=True)
    df.drop(df.loc[(df.Account.isin(['LEND-GACM'])) & (df.Balance<0) & (~df.BalanceType.isin(['Unrealized']))].index, axis=0, inplace=True)
    df.Currency.replace(['USDT_ERC20','SRM_LOCKED','BTCUSD','ETHUSD','EOSUSD','LINKUSD','LTCUSD','ATOM.S','DOT.S','KSM.S','FTM_FANTOM','AUSDC_ETH',
                         'CVXCRV-F','VARIABLEDEBTCRV','BNB_BSC','ZIL_BSC','EUROC_ETH_F5NG','USDTEST','FRXETH','AAVAUSDC','VARIABLEDEBTAVAUSDT',
                         'VARIABLEDEBTUSDT','AURAB-STETH-STABLE-VAULT'],
                        ['USDT','SRM','BTC','ETH','EOS','LINK','LTC','ATOM','DOT','KSM','FTM','AUSDC','CVX','CRV','BNB','ZIL','EUROC','USD','ETH',
                         'AUSDC','USDT','USDT','ETH'], inplace=True)
    df['Notional'] = df.Balance * marks.reindex(df.Currency).fillna(0).Mark.values
    df = _format_fireblocks(df)
    return df

def _format_fireblocks(df):
    temp = df.loc[df.Account=='FBLK',:].copy()
    df.drop(df.loc[df.Account=='FBLK',:].index, axis=0, inplace=True)
    temp = temp.groupby(['Account','Currency','BalanceType','Source','Timestamp','Timestamp_Native']).sum().reset_index()
    df = df.append(temp, ignore_index=True)
    return df

def get_loans(daytime, loans, marks):
    df = _pull_sql_loans(daytime, loans)
    df = _format_loans(daytime, df, marks)
    return df

def _pull_sql_loans(daytime, loans):
    qry = """select a.Account, a.Balance, a.BalanceType, UPPER(a.Currency) as Currency, 
    b.TableName as Source, FORMAT(a.Date_UTC,'yyyyMMdd') as Timestamp, a.AsOf_UTC as Timestamp_Native
    from Operations.balances.EndOfDay_00UTC A LEFT JOIN Operations.balances.sources B on a.Account = b.Account
    where balance != 0 and a.Date_UTC = '{trade_date}' and Currency not like '%SWAP%' and a.Account in {loans}
    and balancetype != 'unrealized' order by Account""".format(trade_date = daytime, loans = tuple(loans))
    df = _get_data(qry)
    return df

def _format_loans(daytime, df, marks):
    df.drop(df.loc[(df.Account.isin(['HUBI-M','HUB2-M'])) & (df.BalanceType!='Margin Loan')].index, axis=0, inplace=True)
    df.drop(df.loc[(df.Account.isin(['WOOX-1-M-E'])) & (df.Balance>=0)].index, axis=0, inplace=True)
    df.drop(df.loc[(df.Account.isin(['FTXE-1-M-E'])) & (df.Balance>=0)].index, axis=0, inplace=True)
    df.drop(df.loc[(df.Account.isin(['OKEX-2-U1'])) & (df.Balance>=0)].index, axis=0, inplace=True)
    df.drop(df.loc[(df.Account.isin(['OKEX-2-S3'])) & (df.Balance>=0)].index, axis=0, inplace=True)
    df.drop(df.loc[(df.Account.isin(['BINE-2-S1-M','BINE-2-S2-M'])) & (df.Balance>=0)].index, axis=0, inplace=True)
    df.drop(df.loc[(df.Account.isin(['LEND-GACM'])) & (df.Balance>=0)].index, axis=0, inplace=True)
    df.Account.replace(['HUBI-M','HUB2-M','WOOX-1-M-E','FTXE-1-M-E','OKEX-2-U1','OKEX-2-S3','BINE-2-S1-M'],
                       ['HUBI-M Margin Loan','HUB2-M Margin Loan','WOOX-1-M-E Margin Loan','FTXE-1-M-E Margin Loan', \
                        'OKEX-2-M Margin Loan','OKEX-2-S3 Margin Loan','BINE-2-S1-M Margin Loan'], inplace=True)
    df['Notional'] = df.Balance * marks.reindex(df.Currency).fillna(0).Mark.values
    return df

def get_edf_cash(daytime, fileloc="/datlib"):

    if daytime.weekday() == 6:
        edf_date = daytime - dt.timedelta(days=2)
    elif daytime.weekday() == 5:
        edf_date = daytime - dt.timedelta(days=1)
    else:
        edf_date = daytime
    
    edf = pd.read_csv('{}/backoffice/EDFman/Downloads/CRYPTO/PFDFMNY_{}.CSV'.format(fileloc, edf_date.strftime('%Y%m%d')))
    edf_cash = edf.T.loc['TOTAL EQUITY','M']

    return edf_cash

def get_bank_balances(daytime):
    qry = "select * from trading.crypto.bankbalances where trade_date = {trade_date}"
    qry = qry.format(qry, trade_date=daytime.strftime('%Y%m%d'))
    b = _get_data(qry)
    
    # Get data for BMO
    tmp = b[b['bfc_name'] == 'BMO']
    bmo = None
    if not tmp.empty:
        bmo = tmp
        bmo.set_index(['trade_date', 'currency'], inplace=True)
    
    # Get data for Silvergate
    tmp = b[b['api_name'] == 'Silvergate']
    slv = None
    if not tmp.empty:
        slv = tmp[['trade_date','balance']].groupby('trade_date').sum()
        slv = slv.loc[daytime.strftime('%Y%m%d'), 'balance']
    
    # Get data for Signature
    tmp = b[b['api_name'] == 'Signature']
    sig = None
    if not tmp.empty:
        sig = tmp[['trade_date','balance']].groupby('trade_date').sum()
        sig = sig.loc[daytime.strftime('%Y%m%d'), 'balance']
    
    return bmo, slv, sig

def summarize_bluescales(td, yd, bs_info):
    # change made by bovas. if there are multiple bluescales in future, use harry's old code for bluescale
    #------------------------------------------------------------------------------------------------------------------------
    tmp = td['assets'].loc[td['assets'].Account.isin(['HUBI-3-S3-F','HUBI-3-S3-M','HUBI-3-S3-E','HUBI-3-S3-P']),['Account','Timestamp','Notional']].append( \
            td['loans'].loc[td['loans'].Account.isin(['HUBI-3-S3-M Loan']),['Account','Timestamp','Notional']])
    bst = {'bluescale6':{'bs_df':tmp}}
    tmp = yd['assets'].loc[yd['assets'].Account.isin(['HUBI-3-S3-F','HUBI-3-S3-M','HUBI-3-S3-E','HUBI-3-S3-P']),['Account','Timestamp','Notional']].append( \
            yd['loans'].loc[yd['loans'].Account.isin(['HUBI-3-S3-M Loan']),['Account','Timestamp','Notional']])
    bsy = {'bluescale6':{'bs_df':tmp}}
    del tmp
    #------------------------------------------------------------------------------------------------------------------------
    
    int_dt_today = td["day"].strftime('%Y%m%d')
    int_dt_yday  = yd["day"].strftime('%Y%m%d')

    total_bluescale_today = 0.0
    total_bluescale_yday  = 0.0

    out = {}
    for k in sorted(bs_info.keys()):
        if bst[k]["bs_df"].empty == True:
            bst[k]["bs_df"].loc[0] = ['HUBI-3-S3-E', int_dt_today, 0]
        if bsy[k]["bs_df"].empty == True:
            bsy[k]["bs_df"].loc[0] = ['HUBI-3-S3-E', int_dt_yday, 0]
        s = bst[k]["bs_df"].append(bsy[k]["bs_df"], ignore_index=True)
        s = s[['Account','Timestamp','Notional']].groupby(['Account','Timestamp']).sum()
        s.reset_index(inplace=True)
        s = s.pivot_table(index=['Account'], values=['Notional'], columns='Timestamp')
        s.fillna(0, axis=1, inplace=True)
        s.columns = [int_dt_yday, int_dt_today]
        s = s.round(0)
        s.sort_index(inplace=True)

        account_list = bs_info[k]["account_list"]

        for a in account_list:
            if a not in s.index.values:
                s.loc[a] = [0, 0]

        # Custom sort by the account_list
        account_order = CategoricalDtype(account_list, ordered=True)
        s['tmp'] = s.index
        s['tmp']=s['tmp'].astype(account_order)
        s = s.sort_values('tmp')
        del s['tmp']

        # Get the specific loan information
        token          = bs_info[k]["loans"][0][0] 
        loan_amount    = bs_info[k]["loans"][0][1] 
        int_rate       = bs_info[k]["loans"][0][2]  
        day_count      = bs_info[k]["loans"][0][3] 
        first_loan_day = bs_info[k]["loans"][0][4]
        interest_paid  = bs_info[k]["interest_paid"]

        # Calculations
        int_days = (td["day"] - first_loan_day).days + 1
        accrued_interest_today = int_days * ( int_rate / day_count) * loan_amount
        accrued_interest_yday  = (int_days - 1) * ( int_rate / day_count) * loan_amount
        accrued_interest_yday  = accrued_interest_yday - interest_paid
        accrued_interest_today = accrued_interest_today - interest_paid

        # Create the summary output
        s.loc['Total BlueScale NAV']     = [s.loc[:,int_dt_yday].sum(), s.loc[:,int_dt_today].sum()]
        s.loc['Less Huobi Contribution'] = [-loan_amount * yd["marks"].loc[token]['Mark'], -loan_amount * td["marks"].loc[token]['Mark']]
        s.loc['Less Accrued Interest']   = [-accrued_interest_yday * yd["marks"].loc[token]['Mark'], -accrued_interest_today * td["marks"].loc[token]['Mark']]
        s.loc['Net BS BFC Assets']       = [s.loc['Total BlueScale NAV', int_dt_yday]  + s.loc['Less Huobi Contribution', int_dt_yday]  + s.loc['Less Accrued Interest', int_dt_yday],
                                            s.loc['Total BlueScale NAV', int_dt_today] + s.loc['Less Huobi Contribution', int_dt_today] + s.loc['Less Accrued Interest', int_dt_today]]
        s['Delta']                       = s.loc[:,int_dt_today] - s.loc[:,int_dt_yday]

        # Compute the summary stats for today
        total_bluescale_today += s.loc['Net BS BFC Assets', int_dt_today]
        total_bluescale_yday  += s.loc['Net BS BFC Assets', int_dt_yday]

        out[k] = s

    return out, total_bluescale_today, total_bluescale_yday

def summarize_all(td, yd, bs):
    
    int_dt_today = td["day"].strftime('%Y%m%d')
    int_dt_yday  = yd["day"].strftime('%Y%m%d')

    # create df
    summary = pd.DataFrame(columns = [int_dt_yday,int_dt_today,'Delta'])
    summary.index.name = 'Account'

    # add values - exchanges
    summary.loc['BTFX-1-M-E',int_dt_yday],summary.loc['BTFX-1-M-E',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='BTFX-1-M-E','Notional'].sum(), td['assets'].loc[td['assets'].Account=='BTFX-1-M-E','Notional'].sum()
    summary.loc['BTSE',int_dt_yday],summary.loc['BTSE',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='BTSE','Notional'].sum(), td['assets'].loc[td['assets'].Account=='BTSE','Notional'].sum()
    summary.loc['FTXE-1-M-E',int_dt_yday],summary.loc['FTXE-1-M-E',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='FTXE-1-M-E','Notional'].sum(), td['assets'].loc[td['assets'].Account=='FTXE-1-M-E','Notional'].sum()
    summary.loc['FUB2-M',int_dt_yday],summary.loc['FUB2-M',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='FUB2-M','Notional'].sum(), td['assets'].loc[td['assets'].Account=='FUB2-M','Notional'].sum()
    summary.loc['FUBI-M',int_dt_yday],summary.loc['FUBI-M',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='FUBI-M','Notional'].sum(), td['assets'].loc[td['assets'].Account=='FUBI-M','Notional'].sum()
    summary.loc['HUB2-E',int_dt_yday],summary.loc['HUB2-E',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='HUB2-E','Notional'].sum(), td['assets'].loc[td['assets'].Account=='HUB2-E','Notional'].sum()
    summary.loc['HUB2-M',int_dt_yday],summary.loc['HUB2-M',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='HUB2-M','Notional'].sum(), td['assets'].loc[td['assets'].Account=='HUB2-M','Notional'].sum()
    summary.loc['HUBI-1-M-P',int_dt_yday],summary.loc['HUBI-1-M-P',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='HUBI-1-M-P','Notional'].sum(), td['assets'].loc[td['assets'].Account=='HUBI-1-M-P','Notional'].sum()
    summary.loc['HUBI-E',int_dt_yday],summary.loc['HUBI-E',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='HUBI-E','Notional'].sum(), td['assets'].loc[td['assets'].Account=='HUBI-E','Notional'].sum()
    summary.loc['HUBI-M',int_dt_yday],summary.loc['HUBI-M',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='HUBI-M','Notional'].sum(), td['assets'].loc[td['assets'].Account=='HUBI-M','Notional'].sum()
    summary.loc['HUBI-1-M-PT',int_dt_yday],summary.loc['HUBI-1-M-PT',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='HUBI-1-M-PT','Notional'].sum(), td['assets'].loc[td['assets'].Account=='HUBI-1-M-PT','Notional'].sum()
    summary.loc['KRKE',int_dt_yday],summary.loc['KRKE',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='KRKE','Notional'].sum(), td['assets'].loc[td['assets'].Account=='KRKE','Notional'].sum()
    summary.loc['LMAC',int_dt_yday],summary.loc['LMAC',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='LMAC','Notional'].sum(), td['assets'].loc[td['assets'].Account=='LMAC','Notional'].sum()
    summary.loc['LMAX',int_dt_yday],summary.loc['LMAX',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='LMAX','Notional'].sum(), td['assets'].loc[td['assets'].Account=='LMAX','Notional'].sum()
    summary.loc['OKEX-2-M',int_dt_yday],summary.loc['OKEX-2-M',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='OKEX-2-M','Notional'].sum(), td['assets'].loc[td['assets'].Account=='OKEX-2-M','Notional'].sum()
    summary.loc['OKEX-2-M-W',int_dt_yday],summary.loc['OKEX-2-M-W',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='OKEX-2-M-W','Notional'].sum(), td['assets'].loc[td['assets'].Account=='OKEX-2-M-W','Notional'].sum()
    summary.loc['OKEX-2-S1',int_dt_yday],summary.loc['OKEX-2-S1',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='OKEX-2-S1','Notional'].sum(), td['assets'].loc[td['assets'].Account=='OKEX-2-S1','Notional'].sum()
    summary.loc['OKEX-2-S2',int_dt_yday],summary.loc['OKEX-2-S2',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='OKEX-2-S2','Notional'].sum(), td['assets'].loc[td['assets'].Account=='OKEX-2-S2','Notional'].sum()
    summary.loc['OKEX-2-S3',int_dt_yday],summary.loc['OKEX-2-S3',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='OKEX-2-S3','Notional'].sum(), td['assets'].loc[td['assets'].Account=='OKEX-2-S3','Notional'].sum()
    summary.loc['PITX-E',int_dt_yday],summary.loc['PITX-E',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='PITX-E','Notional'].sum(), td['assets'].loc[td['assets'].Account=='PITX-E','Notional'].sum()
    summary.loc['WOOX-1-M-E',int_dt_yday],summary.loc['WOOX-1-M-E',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='WOOX-1-M-E','Notional'].sum(), td['assets'].loc[td['assets'].Account=='WOOX-1-M-E','Notional'].sum()
    summary.loc['Fireblocks',int_dt_yday],summary.loc['Fireblocks',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='FBLK','Notional'].sum(), td['assets'].loc[td['assets'].Account=='FBLK','Notional'].sum()
    summary.loc['BULL-1-M-E',int_dt_yday],summary.loc['BULL-1-M-E',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='BULL-1-M-E','Notional'].sum(), td['assets'].loc[td['assets'].Account=='BULL-1-M-E','Notional'].sum()
    summary.loc['BULL-1-M-M',int_dt_yday],summary.loc['BULL-1-M-M',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='BULL-1-M-M','Notional'].sum(), td['assets'].loc[td['assets'].Account=='BULL-1-M-M','Notional'].sum()
    summary.loc['BULL-2-M-E',int_dt_yday],summary.loc['BULL-2-M-E',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='BULL-2-M-E','Notional'].sum(), td['assets'].loc[td['assets'].Account=='BULL-2-M-E','Notional'].sum()
    summary.loc['BULL-2-M-M',int_dt_yday],summary.loc['BULL-2-M-M',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='BULL-2-M-M','Notional'].sum(), td['assets'].loc[td['assets'].Account=='BULL-2-M-M','Notional'].sum()
    summary.loc['BULL-3-M-E',int_dt_yday],summary.loc['BULL-3-M-E',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='BULL-3-M-E','Notional'].sum(), td['assets'].loc[td['assets'].Account=='BULL-3-M-E','Notional'].sum()
    summary.loc['BULL-4-M-E',int_dt_yday],summary.loc['BULL-4-M-E',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='BULL-4-M-E','Notional'].sum(), td['assets'].loc[td['assets'].Account=='BULL-4-M-E','Notional'].sum()
    summary.loc['BINE-MX-S1-P',int_dt_yday],summary.loc['BINE-MX-S1-P',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='BINE-MX-S1-P','Notional'].sum(), td['assets'].loc[td['assets'].Account=='BINE-MX-S1-P','Notional'].sum()
    summary.loc['BINE-MX-S1-F',int_dt_yday],summary.loc['BINE-MX-S1-F',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='BINE-MX-S1-F','Notional'].sum(), td['assets'].loc[td['assets'].Account=='BINE-MX-S1-F','Notional'].sum()
    summary.loc['BINE-MX-S1-E',int_dt_yday],summary.loc['BINE-MX-S1-E',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='BINE-MX-S1-E','Notional'].sum(), td['assets'].loc[td['assets'].Account=='BINE-MX-S1-E','Notional'].sum()
    summary.loc['BINE-MX-S1-M',int_dt_yday],summary.loc['BINE-MX-S1-M',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='BINE-MX-S1-M','Notional'].sum(), td['assets'].loc[td['assets'].Account=='BINE-MX-S1-M','Notional'].sum()
    summary.loc['BINE-2-S1-E',int_dt_yday],summary.loc['BINE-2-S1-E',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='BINE-2-S1-E','Notional'].sum(), td['assets'].loc[td['assets'].Account=='BINE-2-S1-E','Notional'].sum()
    summary.loc['BINE-2-S1-F',int_dt_yday],summary.loc['BINE-2-S1-F',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='BINE-2-S1-F','Notional'].sum(), td['assets'].loc[td['assets'].Account=='BINE-2-S1-F','Notional'].sum()
    summary.loc['BINE-2-S1-M',int_dt_yday],summary.loc['BINE-2-S1-M',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='BINE-2-S1-M','Notional'].sum(), td['assets'].loc[td['assets'].Account=='BINE-2-S1-M','Notional'].sum()
    summary.loc['BINE-2-S1-P',int_dt_yday],summary.loc['BINE-2-S1-P',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='BINE-2-S1-P','Notional'].sum(), td['assets'].loc[td['assets'].Account=='BINE-2-S1-P','Notional'].sum()
    summary.loc['BINE-2-S2-E',int_dt_yday],summary.loc['BINE-2-S2-E',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='BINE-2-S2-E','Notional'].sum(), td['assets'].loc[td['assets'].Account=='BINE-2-S2-E','Notional'].sum()
    summary.loc['BINE-2-S2-F',int_dt_yday],summary.loc['BINE-2-S2-F',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='BINE-2-S2-F','Notional'].sum(), td['assets'].loc[td['assets'].Account=='BINE-2-S2-F','Notional'].sum()
    summary.loc['BINE-2-S2-M',int_dt_yday],summary.loc['BINE-2-S2-M',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='BINE-2-S2-M','Notional'].sum(), td['assets'].loc[td['assets'].Account=='BINE-2-S2-M','Notional'].sum()
    summary.loc['BINE-2-S2-P',int_dt_yday],summary.loc['BINE-2-S2-P',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='BINE-2-S2-P','Notional'].sum(), td['assets'].loc[td['assets'].Account=='BINE-2-S2-P','Notional'].sum()        
    summary.loc['GATE-1-M-C',int_dt_yday],summary.loc['GATE-1-M-C',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='GATE-1-M-C','Notional'].sum(), td['assets'].loc[td['assets'].Account=='GATE-1-M-C','Notional'].sum()
    summary.loc['GATE-1-M-E',int_dt_yday],summary.loc['GATE-1-M-E',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='GATE-1-M-E','Notional'].sum(), td['assets'].loc[td['assets'].Account=='GATE-1-M-E','Notional'].sum()
    summary.loc['GATE-1-M-M',int_dt_yday],summary.loc['GATE-1-M-M',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='GATE-1-M-M','Notional'].sum(), td['assets'].loc[td['assets'].Account=='GATE-1-M-M','Notional'].sum()
    summary.loc['GATE-1-M-PT',int_dt_yday],summary.loc['GATE-1-M-PT',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='GATE-1-M-PT','Notional'].sum(), td['assets'].loc[td['assets'].Account=='GATE-1-M-PT','Notional'].sum()
    summary.loc['DEFI-STRAT-1',int_dt_yday],summary.loc['DEFI-STRAT-1',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='DEFI-STRAT-1','Notional'].sum(), td['assets'].loc[td['assets'].Account=='DEFI-STRAT-1','Notional'].sum()
    summary.loc['DEFI-STRAT-2',int_dt_yday],summary.loc['DEFI-STRAT-2',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='DEFI-STRAT-2','Notional'].sum(), td['assets'].loc[td['assets'].Account=='DEFI-STRAT-2','Notional'].sum()
    summary.loc['DEFI-STRAT-3',int_dt_yday],summary.loc['DEFI-STRAT-3',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='DEFI-STRAT-3','Notional'].sum(), td['assets'].loc[td['assets'].Account=='DEFI-STRAT-3','Notional'].sum()
    summary.loc['DEFI-STRAT-4',int_dt_yday],summary.loc['DEFI-STRAT-4',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='DEFI-STRAT-4','Notional'].sum(), td['assets'].loc[td['assets'].Account=='DEFI-STRAT-4','Notional'].sum()
    summary.loc['DEFI-STRAT-5',int_dt_yday],summary.loc['DEFI-STRAT-5',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='DEFI-STRAT-5','Notional'].sum(), td['assets'].loc[td['assets'].Account=='DEFI-STRAT-5','Notional'].sum()
    summary.loc['DEFI-STRAT-6',int_dt_yday],summary.loc['DEFI-STRAT-6',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='DEFI-STRAT-6','Notional'].sum(), td['assets'].loc[td['assets'].Account=='DEFI-STRAT-6','Notional'].sum()
    summary.loc['DEFI-STRAT-8',int_dt_yday],summary.loc['DEFI-STRAT-8',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='DEFI-STRAT-8','Notional'].sum(), td['assets'].loc[td['assets'].Account=='DEFI-STRAT-8','Notional'].sum()
    summary.loc['DYDX-1-M-P',int_dt_yday],summary.loc['DYDX-1-M-P',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='DYDX-1-M-P','Notional'].sum(), td['assets'].loc[td['assets'].Account=='DYDX-1-M-P','Notional'].sum()
    
    # total exchanges
    summary.loc['Total_Exchanges',int_dt_yday],summary.loc['Total_Exchanges',int_dt_today] = \
    summary.loc[:,int_dt_yday].sum(), summary.loc[:,int_dt_today].sum()

    # add values - non-exchange assets
    summary.loc['EDF',int_dt_yday],summary.loc['EDF',int_dt_today] = \
    yd['edf'], td['edf']

    # total balances at clearing brokers
    summary.loc['Total Balances at Clearing Brokers',int_dt_yday],summary.loc['Total Balances at Clearing Brokers',int_dt_today] = \
    summary.loc['EDF',int_dt_yday], summary.loc['EDF',int_dt_today]

    summary.loc['Silvergate',int_dt_yday],summary.loc['Silvergate',int_dt_today] = \
    yd['slv'], td['slv']
    summary.loc['Signature',int_dt_yday],summary.loc['Signature',int_dt_today] = \
    yd['sig'], td['sig']
    summary.loc['BMO_USD',int_dt_yday],summary.loc['BMO_USD',int_dt_today] = \
    yd['bmo'].loc[(int_dt_yday,'USD'),'balance'], td['bmo'].loc[(int_dt_today,'USD'),'balance']
    summary.loc['BMO_EUR',int_dt_yday],summary.loc['BMO_EUR',int_dt_today] = \
    yd['bmo'].loc[(int_dt_yday,'EUR'),'balance'] * yd['marks'].loc['EUR'].Mark, td['bmo'].loc[(int_dt_today,'EUR'),'balance'] * td['marks'].loc['EUR'].Mark

    # total cash in banks
    # when deleting FIRST/ LAST loan below, update 'Total Cash in Banks' index as well
    summary.loc['Total Cash in Banks',int_dt_yday],summary.loc['Total Cash in Banks',int_dt_today] = \
    summary.loc['Silvergate':'BMO_EUR',int_dt_yday].sum(), \
    summary.loc['Silvergate':'BMO_EUR',int_dt_today].sum()

    summary.loc['SHFT-1-W',int_dt_yday],summary.loc['SHFT-1-W',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='SHFT-1-W','Notional'].sum(), td['assets'].loc[td['assets'].Account=='SHFT-1-W','Notional'].sum()
    
    summary.loc['Wedbush',int_dt_yday],summary.loc['Wedbush',int_dt_today] = 79941.92, 79941.92

    # total assets
    summary.loc['Total Assets',int_dt_yday],summary.loc['Total Assets',int_dt_today] = \
    summary.loc[:,int_dt_yday].sum()-summary.loc['Total_Exchanges',int_dt_yday]-summary.loc['Total Balances at Clearing Brokers',int_dt_yday]- \
        summary.loc['Total Cash in Banks',int_dt_yday], summary.loc[:,int_dt_today].sum()-summary.loc['Total_Exchanges',int_dt_today]- \
        summary.loc['Total Balances at Clearing Brokers',int_dt_today]-summary.loc['Total Cash in Banks',int_dt_today]

    # add loans
    # when deleting FIRST/ LAST loan below, update 'Total Loan' index as well
    summary.loc['HUB2-M Margin Loan',int_dt_yday],summary.loc['HUB2-M Margin Loan',int_dt_today] = \
    yd['loans'].loc[yd['loans'].Account=='HUB2-M Margin Loan','Notional'].sum(), td['loans'].loc[td['loans'].Account=='HUB2-M Margin Loan','Notional'].sum()
    summary.loc['HUBI-M Margin Loan',int_dt_yday],summary.loc['HUBI-M Margin Loan',int_dt_today] = \
    yd['loans'].loc[yd['loans'].Account=='HUBI-M Margin Loan','Notional'].sum(), td['loans'].loc[td['loans'].Account=='HUBI-M Margin Loan','Notional'].sum()
    summary.loc['FTXE-1-M-E Margin Loan',int_dt_yday],summary.loc['FTXE-1-M-E Margin Loan',int_dt_today] = \
    yd['loans'].loc[yd['loans'].Account=='FTXE-1-M-E Margin Loan','Notional'].sum(), td['loans'].loc[td['loans'].Account=='FTXE-1-M-E Margin Loan','Notional'].sum()
    summary.loc['LEND-GACM',int_dt_yday],summary.loc['LEND-GACM',int_dt_today] = \
    yd['loans'].loc[yd['loans'].Account=='LEND-GACM','Notional'].sum() + yd['assets'].loc[yd['assets'].Account=='LEND-GACM','Notional'].sum(), \
    td['loans'].loc[td['loans'].Account=='LEND-GACM','Notional'].sum() + td['assets'].loc[td['assets'].Account=='LEND-GACM','Notional'].sum()
    summary.loc['LEND-HUBI',int_dt_yday],summary.loc['LEND-HUBI',int_dt_today] = \
    yd['loans'].loc[yd['loans'].Account=='LEND-HUBI','Notional'].sum(), td['loans'].loc[td['loans'].Account=='LEND-HUBI','Notional'].sum()
    summary.loc['LEND-LMAC',int_dt_yday],summary.loc['LEND-LMAC',int_dt_today] = \
    yd['loans'].loc[yd['loans'].Account=='LEND-LMAC','Notional'].sum(), td['loans'].loc[td['loans'].Account=='LEND-LMAC','Notional'].sum()
    summary.loc['LEND-OKEX',int_dt_yday],summary.loc['LEND-OKEX',int_dt_today] = \
    yd['loans'].loc[yd['loans'].Account=='LEND-OKEX','Notional'].sum(), td['loans'].loc[td['loans'].Account=='LEND-OKEX','Notional'].sum()
    summary.loc['LEND-OXTF',int_dt_yday],summary.loc['LEND-OXTF',int_dt_today] = \
    yd['loans'].loc[yd['loans'].Account=='LEND-OXTF','Notional'].sum(), td['loans'].loc[td['loans'].Account=='LEND-OXTF','Notional'].sum()
    summary.loc['LEND-PITX',int_dt_yday],summary.loc['LEND-PITX',int_dt_today] = \
    yd['loans'].loc[yd['loans'].Account=='LEND-PITX','Notional'].sum(), td['loans'].loc[td['loans'].Account=='LEND-PITX','Notional'].sum()
    summary.loc['LEND-XRPF',int_dt_yday],summary.loc['LEND-XRPF',int_dt_today] = \
    yd['loans'].loc[yd['loans'].Account=='LEND-XRPF','Notional'].sum(), td['loans'].loc[td['loans'].Account=='LEND-XRPF','Notional'].sum()
    summary.loc['LEND-BULL',int_dt_yday],summary.loc['LEND-BULL',int_dt_today] = \
    yd['loans'].loc[yd['loans'].Account=='LEND-BULL','Notional'].sum(), td['loans'].loc[td['loans'].Account=='LEND-BULL','Notional'].sum()
    summary.loc['OKEX-2-M Margin Loan',int_dt_yday],summary.loc['OKEX-2-M Margin Loan',int_dt_today] = \
    yd['loans'].loc[yd['loans'].Account=='OKEX-2-M Margin Loan','Notional'].sum(), td['loans'].loc[td['loans'].Account=='OKEX-2-M Margin Loan','Notional'].sum()    
    summary.loc['OKEX-2-S3 Margin Loan',int_dt_yday],summary.loc['OKEX-2-S3 Margin Loan',int_dt_today] = \
    yd['loans'].loc[yd['loans'].Account=='OKEX-2-S3 Margin Loan','Notional'].sum(), td['loans'].loc[td['loans'].Account=='OKEX-2-S3 Margin Loan','Notional'].sum()
    summary.loc['BINE-2-S1-M Margin Loan',int_dt_yday],summary.loc['BINE-2-S1-M Margin Loan',int_dt_today] = \
    yd['loans'].loc[yd['loans'].Account=='BINE-2-S1-M Margin Loan','Notional'].sum(), td['loans'].loc[td['loans'].Account=='BINE-2-S1-M Margin Loan','Notional'].sum()
    summary.loc['BINE-2-S2-M Margin Loan',int_dt_yday],summary.loc['BINE-2-S2-M Margin Loan',int_dt_today] = \
    yd['loans'].loc[yd['loans'].Account=='BINE-2-S2-M Margin Loan','Notional'].sum(), td['loans'].loc[td['loans'].Account=='BINE-2-S2-M Margin Loan','Notional'].sum()
    summary.loc['WOOX-1-M-E Margin Loan',int_dt_yday],summary.loc['WOOX-1-M-E Margin Loan',int_dt_today] = \
    yd['loans'].loc[yd['loans'].Account=='WOOX-1-M-E Margin Loan','Notional'].sum(), td['loans'].loc[td['loans'].Account=='WOOX-1-M-E Margin Loan','Notional'].sum()
    summary.loc['BULL-1-M-M Loan',int_dt_yday],summary.loc['BULL-1-M-M Loan',int_dt_today] = \
    yd['loans'].loc[yd['loans'].Account=='BULL-1-M-M Loan','Notional'].sum(), td['loans'].loc[td['loans'].Account=='BULL-1-M-M Loan','Notional'].sum()
    summary.loc['BULL-2-M-M Loan',int_dt_yday],summary.loc['BULL-2-M-M Loan',int_dt_today] = \
    yd['loans'].loc[yd['loans'].Account=='BULL-2-M-M Loan','Notional'].sum(), td['loans'].loc[td['loans'].Account=='BULL-2-M-M Loan','Notional'].sum()
    summary.loc['GATE-1-M-M Loan',int_dt_yday],summary.loc['GATE-1-M-M Loan',int_dt_today] = \
    yd['loans'].loc[yd['loans'].Account=='GATE-1-M-M Loan','Notional'].sum(), td['loans'].loc[td['loans'].Account=='GATE-1-M-M Loan','Notional'].sum()
    summary.loc['DEFI-STRAT-1 Loan',int_dt_yday],summary.loc['DEFI-STRAT-1 Loan',int_dt_today] = \
    yd['loans'].loc[yd['loans'].Account=='DEFI-STRAT-1 Loan','Notional'].sum(), td['loans'].loc[td['loans'].Account=='DEFI-STRAT-1 Loan','Notional'].sum()
    summary.loc['DEFI-STRAT-8 Loan',int_dt_yday],summary.loc['DEFI-STRAT-8 Loan',int_dt_today] = \
    yd['loans'].loc[yd['loans'].Account=='DEFI-STRAT-8 Loan','Notional'].sum(), td['loans'].loc[td['loans'].Account=='DEFI-STRAT-8 Loan','Notional'].sum()

    # Total Loan
    summary.loc['Total Loan',int_dt_yday],summary.loc['Total Loan',int_dt_today] = \
    summary.loc['HUB2-M Margin Loan':'DEFI-STRAT-8 Loan',int_dt_yday].sum(), \
    summary.loc['HUB2-M Margin Loan':'DEFI-STRAT-8 Loan',int_dt_today].sum()

    # Bluescales
    summary.loc['HUBI-3-S3-F',int_dt_yday],summary.loc['HUBI-3-S3-F',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='HUBI-3-S3-F','Notional'].sum(), td['assets'].loc[td['assets'].Account=='HUBI-3-S3-F','Notional'].sum()
    summary.loc['HUBI-3-S3-M',int_dt_yday],summary.loc['HUBI-3-S3-M',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='HUBI-3-S3-M','Notional'].sum(), td['assets'].loc[td['assets'].Account=='HUBI-3-S3-M','Notional'].sum()
    summary.loc['HUBI-3-S3-M Loan',int_dt_yday],summary.loc['HUBI-3-S3-M Loan',int_dt_today] = \
    yd['loans'].loc[yd['loans'].Account=='HUBI-3-S3-M Loan','Notional'].sum(), td['loans'].loc[td['loans'].Account=='HUBI-3-S3-M Loan','Notional'].sum()
    summary.loc['HUBI-3-S3-P',int_dt_yday],summary.loc['HUBI-3-S3-P',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='HUBI-3-S3-P','Notional'].sum(), td['assets'].loc[td['assets'].Account=='HUBI-3-S3-P','Notional'].sum()
    summary.loc['HUBI-3-S3-E',int_dt_yday],summary.loc['HUBI-3-S3-E',int_dt_today] = \
    yd['assets'].loc[yd['assets'].Account=='HUBI-3-S3-E','Notional'].sum(), td['assets'].loc[td['assets'].Account=='HUBI-3-S3-E','Notional'].sum()
    # Total BlueScale NAV
    summary.loc['Total BlueScale NAV',int_dt_yday],summary.loc['Total BlueScale NAV',int_dt_today] = \
    summary.loc['HUBI-3-S3-F':'HUBI-3-S3-E',int_dt_yday].sum(),summary.loc['HUBI-3-S3-F':'HUBI-3-S3-E',int_dt_today].sum()
    # Net BS BFC Assets
    summary = summary.append(bs[0]['bluescale6'].loc['Less Huobi Contribution':'Net BS BFC Assets',])
    summary.loc['',int_dt_yday],summary.loc['',int_dt_today] = '',''

    # Net Assets
    summary.loc['Net Assets',int_dt_yday],summary.loc['Net Assets',int_dt_today] = \
    summary.loc[['Total Assets','Total Loan','Net BS BFC Assets'],int_dt_yday].sum(), \
    summary.loc[['Total Assets','Total Loan','Net BS BFC Assets'],int_dt_today].sum()
    # calculate delta
    summary.loc[summary.index!='','Delta'] = summary.loc[summary.index!='',int_dt_today].values - summary.loc[summary.index!='',int_dt_yday].values
    summary.fillna('', inplace=True)
    
    return summary

# def export_to_excel(td, yd, summary, fileloc):
#     data_dict = _create_dict_for_export(td, yd, summary)
#     xlname = '{}/Operations/EOD Asset Report/{DATE}_EODAssets_00UTC.xlsx'.format(fileloc, DATE=td["day"].strftime('%Y%m%d'))
#     xlname = 'C:/Users/bpr/Desktop/{DATE}_EODAssets_00UTC.xlsx'.format(fileloc, DATE=td["day"].strftime('%Y%m%d'))
#     Excelwriter = pd.ExcelWriter(xlname,engine="xlsxwriter")
#     for n, key in enumerate (data_dict):
#         data_dict[key] = data_dict[key].reset_index()
#         data_dict[key] = data_dict[key].T.reset_index(drop=False).T
#         data_dict[key].to_excel(Excelwriter, sheet_name=key, index=False, header=False)
#     wb = Excelwriter.book
#     sh = Excelwriter.sheets['Summary']
#     number_format = wb.add_format({'num_format': '#,##0'})
#     sh.set_column('A:A', 29)
#     sh.set_column('B:D', 11.5, number_format)
#     Excelwriter.save()
#     Excelwriter.close()

def _create_dict_for_export(td, yd, summary):
    dict_export = {'Summary':summary, 'Today_Asset_Detail': td['assets'], 'Yesterday_Asset_Detail': yd['assets'], 
                   'Today_Loan_Detail': td['loans'], 'Yesterday_Loan_Detail': yd['loans'], 'Today_Prices': td['marks'], 'Yesterday_Prices': yd['marks']}
    dict_export['Summary_Exchange_Balances_00UTC'] = dict_export['Today_Asset_Detail'].append(dict_export['Today_Loan_Detail'], ignore_index=False).set_index('Account')
    return dict_export

def export_to_excel(td, yd, summary, fileloc):
    data_dict = _create_dict_for_export(td, yd, summary)
    wb = load_workbook(filename = r"{}/Operations/EOD Asset Report/NAV REPORT 00UTC.xlsx".format(fileloc))
    for k, v in data_dict.items():
        ws = wb[k]
        df = v.reset_index()
        rows = dataframe_to_rows(df, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                 ws.cell(row=r_idx, column=c_idx, value=value)
    ws = wb['NAV Summary 00UTC']
    ws['D6'] = "{} {}, {} 00:00UTC".format(td["day"].strftime('%B'), td["day"].day, td["day"].year)
    wb.save(r"{}/Operations/EOD Asset Report/{}_NAV REPORT 00UTC.xlsx".format(fileloc, td["day"].strftime('%Y%m%d')))
    wb.save(r"C:/Users/bpr/Documents/b.works/pending_navs/00utc/sent/{}_NAV REPORT 00UTC.xlsx".format(td["day"].strftime('%Y%m%d')))